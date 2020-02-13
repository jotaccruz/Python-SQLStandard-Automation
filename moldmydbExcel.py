from openpyxl import Workbook, cell
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, NamedStyle,Font, Color, Border, Side, \
Alignment
from openpyxl.styles.differential import DifferentialStyle
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
#from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

def as_text(value):
    if value is None:
        return ""
    return str(value)

def xlsxGen(excelf,section,tree,tree_name,tree_dic):
    #Style
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(fill=red_fill)
    r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    r.formula = ['$B1="Missing"','$B1="Disabled"']

    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")

    #-----

    wb=excelf
    ws = wb.active
    if ("Sheet" in wb.sheetnames):
        ws.title = section
    else:
        if section in wb.sheetnames:
            ws = wb[section]
        else:
            ws = wb.create_sheet(title=section)

    columnslist=tree['columns']
    row=1
    col=1
    for titles in columnslist:
        ws.cell (row=row,column=col,value=titles)
        ws.cell (row,col).border = Border(top=thin, left=thin, right=thin, \
        bottom=thin)
        col=col+1

    row=3
    for lists in tree.get_children():
        col=1
        for values in tree.item(lists)['values']:
            ws.cell (column=col,row=row,value=values)
            ws.conditional_formatting.add("A1:C500", r)
            col=col+1
        row=row+1

    for key in tree_dic:
        #print (key)
        if key==tree_name:
            i=1
            for values in tree_dic[key]:
                #print (get_column_letter(i))
                ws.column_dimensions[get_column_letter(i)].width = values
                i=i+1

        #print(titles)

    #ws['A1']="Hola"
    #wb.save('test.xlsx')

    #Applying Style
    return wb
